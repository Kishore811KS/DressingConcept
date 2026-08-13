from flask import Blueprint, request, jsonify, make_response
from app.models.billing import Bill, BillItem, Payment
from app.models.product import Product
from app.models.current_company import Company
from app.models.customer_rewards import CustomerRewards
from app.models.sale_return import SaleReturn
from app import db
from sqlalchemy import or_, and_, func, text
from datetime import datetime, timedelta
import traceback
import random
import string
from dateutil.relativedelta import relativedelta  # Add this import for warranty calculation

billing_bp = Blueprint("billing_bp", __name__)

def get_fy_letter(dt=None):
    """
    Returns letter prefix for financial year.
    FY 2026-2027 (April 1, 2026 to March 31, 2027) -> 'A'
    FY 2027-2028 (April 1, 2027 to March 31, 2028) -> 'B'
    FY 2028-2029 (April 1, 2028 to March 31, 2029) -> 'C'
    and so on.
    """
    if dt is None:
        dt = datetime.now()
    year = dt.year
    if dt.month >= 4:
        start_yr = year
    else:
        start_yr = year - 1
    
    base_year = 2026
    offset = start_yr - base_year
    if offset < 0:
        offset = 0
    
    if offset < 26:
        return chr(65 + offset)
    else:
        first = chr(65 + (offset // 26) - 1)
        second = chr(65 + (offset % 26))
        return f"{first}{second}"


def get_current_financial_year(dt=None):
    """
    Returns financial year string e.g. '26-27' for April 1, 2026 to March 31, 2027.
    Budget year runs from April to March.
    """
    if dt is None:
        dt = datetime.now()
    year = dt.year
    if dt.month >= 4:
        start_yr = year % 100
        end_yr = (year + 1) % 100
    else:
        start_yr = (year - 1) % 100
        end_yr = year % 100
    return f"{start_yr:02d}-{end_yr:02d}"


def generate_unique_bill_number(bill_type='N', dt=None):
    """
    Generate unique bill number with letter prefix per FY and 4-digit sequence.
    Format: A0001N, A0002N... for original bills
            A0001R, A0002R... for sales returns
    Next FY (B): B0001N, B0001R...
    """
    bill_type = str(bill_type).strip().upper()
    if bill_type not in ['N', 'R']:
        bill_type = 'N'
    
    fy_letter = get_fy_letter(dt)
    import re
    from app.models.sale_return import SaleReturn
    
    max_seq = 0
    pattern = re.compile(rf'^{fy_letter}(\d+){bill_type}$', re.IGNORECASE)
    
    if bill_type == 'N':
        bills = Bill.query.all()
        for b in bills:
            if not b.bill_number:
                continue
            b_str = str(b.bill_number).strip().upper()
            m = pattern.match(b_str)
            if m:
                num = int(m.group(1))
                if num > max_seq:
                    max_seq = num
            else:
                fy_prefix = get_current_financial_year(dt)
                if b_str.startswith(f"{fy_prefix}/"):
                    seq_part = b_str.split('/', 1)[1]
                    digits = re.findall(r'\d+', seq_part)
                    if digits:
                        num = int(digits[0])
                        if num > max_seq:
                            max_seq = num
    else:
        returns = SaleReturn.query.all()
        for r in returns:
            if not r.return_number:
                continue
            r_str = str(r.return_number).strip().upper()
            m = pattern.match(r_str)
            if m:
                num = int(m.group(1))
                if num > max_seq:
                    max_seq = num
            else:
                fy_prefix = get_current_financial_year(dt)
                if r_str.startswith(f"{fy_prefix}/"):
                    seq_part = r_str.split('/', 1)[1]
                    digits = re.findall(r'\d+', seq_part)
                    if digits:
                        num = int(digits[0])
                        if num > max_seq:
                            max_seq = num

    next_num = max_seq + 1
    candidate = f"{fy_letter}{next_num:04d}{bill_type}"

    if bill_type == 'N':
        while Bill.query.filter_by(bill_number=candidate).first():
            next_num += 1
            candidate = f"{fy_letter}{next_num:04d}{bill_type}"
    else:
        while SaleReturn.query.filter_by(return_number=candidate).first():
            next_num += 1
            candidate = f"{fy_letter}{next_num:04d}{bill_type}"

    return candidate


@billing_bp.route("/billing/next-bill-number", methods=["GET"])
def get_next_bill_number_route():
    """Get the next sequential bill number for N (normal) or R (return) mode."""
    try:
        bill_type = request.args.get('type', 'N').strip().upper()
        if bill_type not in ['N', 'R']:
            bill_type = 'N'
        next_num = generate_unique_bill_number(bill_type)
        # Short display bill number for POS receipt e.g. "1N" or "1R"
        display_num = next_num.split('/')[-1] if '/' in next_num else next_num
        return jsonify({"nextBillNumber": display_num, "fullBillNumber": next_num, "displayBillNumber": display_num, "type": bill_type}), 200
    except Exception as e:
        print(f"Error generating next bill number: {str(e)}")
        return jsonify({"error": str(e)}), 400


@billing_bp.route("/billing/bills/return-details/<path:bill_number>", methods=["GET"])
def get_bill_return_details(bill_number):
    """
    Get bill details along with cumulative prior return quantities for Sales Return mode.
    Only products with remaining quantity > 0 are returnable.
    """
    try:
        from app.models.sale_return import SaleReturn
        clean_no = str(bill_number).strip()
        bill = Bill.query.filter_by(bill_number=clean_no).first()
        if not bill:
            fy_prefix = get_current_financial_year()
            raw_seq = clean_no.split('/')[-1].rstrip('N').rstrip('n').rstrip('R').rstrip('r')
            if raw_seq.isdigit():
                num = int(raw_seq)
                possible = [
                    f"{fy_prefix}/{num}N",
                    f"{fy_prefix}/{num}R",
                    f"{fy_prefix}/{raw_seq.zfill(4)}N",
                    f"{num}N",
                    f"{raw_seq.zfill(4)}N",
                    clean_no
                ]
                bill = Bill.query.filter(Bill.bill_number.in_(possible)).first()
            
        if not bill:
            return jsonify({"error": f"Bill #{bill_number} not found"}), 404

        prior_returns = SaleReturn.query.filter_by(original_bill_number=bill.bill_number).all()
        already_returned = {}
        for pr in prior_returns:
            for item in pr.items:
                k_code = str(item.product_code or '').strip().lower()
                k_name = str(item.product_name or '').strip().lower()
                qty = item.returned_quantity or 0
                if k_code:
                    already_returned[k_code] = already_returned.get(k_code, 0) + qty
                if k_name:
                    already_returned[k_name] = already_returned.get(k_name, 0) + qty

        items_details = []
        for item in bill.items:
            k_code = str(item.product_code or '').strip().lower()
            k_name = str(item.product_name or '').strip().lower()
            prev_ret = max(already_returned.get(k_code, 0), already_returned.get(k_name, 0))
            orig_qty = item.quantity or 1
            rem_qty = max(0, orig_qty - prev_ret)

            items_details.append({
                'id': item.id,
                'productId': item.product_id,
                'productCode': item.product_code or '',
                'productName': item.product_name or '',
                'productModel': item.product_model or '',
                'productType': item.product_type or '',
                'tax': item.tax or 5.0,
                'unit': 'PCS',
                'sellPrice': item.sell_price or 0,
                'mrp': item.sell_price or 0,
                'originalQuantity': orig_qty,
                'alreadyReturnedQuantity': prev_ret,
                'remainingQuantity': rem_qty,
                'total': item.total or 0
            })

        return jsonify({
            'success': True,
            'billNumber': bill.bill_number,
            'customerName': bill.customer_name,
            'customerPhone': bill.customer_phone,
            'customerAddress': bill.customer_address,
            'customerEmail': bill.customer_email,
            'createdByName': bill.created_by_name,
            'createdAt': bill.created_at.isoformat() if bill.created_at else None,
            'items': items_details
        }), 200

    except Exception as e:
        print(f"Error fetching return details: {str(e)}")
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 400



# ------------------ SEARCH PRODUCTS FOR BILLING ------------------
@billing_bp.route("/billing/search-products", methods=["GET"])
def search_products_for_billing():
    """Search products by name, model, or type for billing"""
    try:
        query = request.args.get('q', '').strip()
        
        if not query or len(query) < 2:
            return jsonify([]), 200
            
        # Search in name, model, and type, only show products with stock > 0
        products = Product.query.filter(
            or_(
                Product.product_code.ilike(f'%{query}%'),
                Product.name.ilike(f'%{query}%'),
                Product.description.ilike(f'%{query}%'),
                Product.model.ilike(f'%{query}%'),
                Product.type.ilike(f'%{query}%')
            )
        ).filter(Product.quantity > 0).limit(10).all()
        
        result = [{
            'id': p.id,
            'productCode': p.product_code or '',
            'name': p.name,
            'description': p.description or '',
            'model': p.model or '',
            'unit': p.unit or 'PCS',
            'tax': p.tax or 0,
            'mrp': p.mrp or p.buy_price or p.sell_price,
            'discountPercent': p.discount_percent or 0,
            'discountAmount': p.discount_amount or 0,
            'netPrice': p.net_price or p.sell_price,
            'salesPerson': p.sales_person or '',
            'type': p.type or '',
            'sellPrice': p.sell_price,
            'quantity': p.quantity,
            'inStock': p.quantity > 0
        } for p in products]
        
        return jsonify(result), 200
        
    except Exception as e:
        print(f"Search error: {str(e)}")
        print(traceback.format_exc())
        return jsonify({"error": "Failed to search products"}), 400


# ------------------ GET PRODUCT BY BARCODE ------------------
@billing_bp.route("/billing/product/barcode/<string:barcode>", methods=["GET"])
def get_product_by_barcode(barcode):
    """Get product by barcode for quick billing"""
    try:
        if not barcode:
            return jsonify({"error": "Barcode is required"}), 400
            
        product = Product.query.filter_by(barcode=barcode).first()
        
        if not product:
            return jsonify({"error": "Product not found"}), 404
            
        if product.quantity <= 0:
            return jsonify({"error": "Product out of stock"}), 400
            
        return jsonify({
            'id': product.id,
            'name': product.name,
            'model': product.model or '',
            'type': product.type or '',
            'sellPrice': product.sell_price,
            'quantity': product.quantity
        }), 200
        
    except Exception as e:
        print(f"Barcode error: {str(e)}")
        return jsonify({"error": "Failed to fetch product"}), 400


# ------------------ GET CUSTOMER BY PHONE NUMBER ------------------
@billing_bp.route("/billing/customer/<string:phone_number>", methods=["GET"])
def get_customer_by_phone(phone_number):
    """Get customer details by phone number to check for duplicates"""
    try:
        if not phone_number:
            return jsonify({"error": "Phone number is required"}), 400
        
        # Find existing rewards
        rewards = CustomerRewards.query.filter_by(phone=phone_number).first()
        
        # Find existing bills with this phone number (get the most recent)
        existing_customer = Bill.query.filter_by(customer_phone=phone_number).order_by(Bill.created_at.desc()).first()
        
        if existing_customer or rewards:
            return jsonify({
                'exists': True,
                'customer': {
                    'name': rewards.name if rewards and rewards.name else (existing_customer.customer_name if existing_customer else 'Walk-in Customer'),
                    'phone': phone_number,
                    'email': existing_customer.customer_email if existing_customer else '',
                    'gst': existing_customer.customer_gst if existing_customer else '',
                    'address': existing_customer.customer_address if existing_customer else '',
                    'type': existing_customer.customer_type if existing_customer else 'regular',
                    'reward_points': rewards.current_balance if rewards else 0.0
                }
            }), 200
        else:
            return jsonify({
                'exists': False,
                'customer': None
            }), 200
            
    except Exception as e:
        print(f"Get customer error: {str(e)}")
        return jsonify({"error": "Failed to fetch customer details"}), 400


# ------------------ GET ALL REWARDS ------------------
@billing_bp.route("/billing/rewards", methods=["GET"])
def get_all_rewards():
    """Get all customer rewards mapping phone to points"""
    try:
        rewards = CustomerRewards.query.all()
        return jsonify({
            r.phone: r.current_balance for r in rewards
        }), 200
    except Exception as e:
        print(f"Get rewards error: {str(e)}")
        return jsonify({"error": "Failed to fetch rewards"}), 400


# ------------------ GET ALL CUSTOMERS (consolidated) ------------------
@billing_bp.route("/billing/customers", methods=["GET"])
def get_all_customers():
    """Get unique customers consolidated from CustomerRewards and bill history"""
    try:
        rewards = CustomerRewards.query.all()
        from app.models.sale_return import SaleReturn
        returns_query = db.session.query(
            SaleReturn.customer_phone,
            func.sum(SaleReturn.total_return_amount).label('total_return')
        ).filter(SaleReturn.customer_phone.isnot(None), SaleReturn.customer_phone != '')\
         .group_by(SaleReturn.customer_phone).all()
        returns_map = {r[0]: float(r[1] or 0.0) for r in returns_query if r[0]}

        used_query = db.session.query(
            Bill.customer_phone,
            func.sum(getattr(Bill, 'sales_return_amount', 0)).label('total_used')
        ).filter(Bill.customer_phone.isnot(None), Bill.customer_phone != '')\
         .group_by(Bill.customer_phone).all()
        used_map = {u[0]: float(u[1] or 0.0) for u in used_query if u[0]}

        all_phones = set(list(returns_map.keys()) + list(used_map.keys()))
        returns_dict = {
            p: max(0.0, round(returns_map.get(p, 0.0) - used_map.get(p, 0.0), 2))
            for p in all_phones
        }

        customers_dict = {}

        for r in rewards:
            if r.phone:
                customers_dict[r.phone] = {
                    'name': r.name or 'Walk-in Customer',
                    'firstName': r.first_name or '',
                    'lastName': r.last_name or '',
                    'phone': r.phone,
                    'email': r.email or '',
                    'gst': r.gst or '',
                    'address': r.address or '',
                    'type': r.customer_type or 'regular',
                    'dateOfBirth': r.date_of_birth.isoformat() if r.date_of_birth else '',
                    'memberId': r.member_id or '',
                    'weddingAnniversary': r.wedding_anniversary.isoformat() if r.wedding_anniversary else '',
                    'celebrationDate': r.celebration_date.isoformat() if r.celebration_date else '',
                    'isClassicCustomer': bool(r.is_classic_customer),
                    'isSupplier': bool(r.is_supplier),
                    'supplierIGST': round(r.supplier_igst or 0, 2),
                    'billCount': r.bill_count or 0,
                    'rewardPoints': r.current_balance or 0,
                    'totalSpent': r.total_spend or 0,
                    'salesReturnAmount': returns_dict.get(r.phone, 0.0),
                    'lastVisit': r.updated_at.isoformat() if r.updated_at else (r.created_at.isoformat() if r.created_at else None)
                }

        bills_query = db.session.query(
            Bill.customer_name,
            Bill.customer_phone,
            Bill.customer_email,
            Bill.customer_gst,
            Bill.customer_address,
            Bill.customer_type,
            func.count(Bill.id).label('bill_count'),
            func.max(Bill.created_at).label('last_visit')
        ).filter(Bill.customer_phone.isnot(None), Bill.customer_phone != '')\
         .group_by(Bill.customer_name, Bill.customer_phone, Bill.customer_email,
                   Bill.customer_gst, Bill.customer_address, Bill.customer_type).all()

        for c in bills_query:
            phone = c[1]
            if phone:
                if phone in customers_dict:
                    cust = customers_dict[phone]
                    if c[0] and c[0] != 'Walk-in Customer':
                        cust['name'] = c[0]
                    if c[2] and not cust['email']: cust['email'] = c[2]
                    if c[3] and not cust['gst']:   cust['gst'] = c[3]
                    if c[4] and not cust['address']: cust['address'] = c[4]
                    if c[5]: cust['type'] = c[5]
                    if c[6]: cust['billCount'] = c[6]
                    if c[7]: cust['lastVisit'] = c[7].isoformat()
                    if phone in returns_dict: cust['salesReturnAmount'] = returns_dict[phone]
                else:
                    customers_dict[phone] = {
                        'name': c[0] or 'Walk-in Customer',
                        'firstName': '',
                        'lastName': '',
                        'phone': phone,
                        'email': c[2] or '',
                        'gst': c[3] or '',
                        'address': c[4] or '',
                        'type': c[5] or 'regular',
                        'dateOfBirth': '',
                        'memberId': '',
                        'weddingAnniversary': '',
                        'celebrationDate': '',
                        'isClassicCustomer': False,
                        'isSupplier': False,
                        'supplierIGST': 0.0,
                        'billCount': c[6] or 0,
                        'rewardPoints': 0,
                        'totalSpent': 0,
                        'salesReturnAmount': returns_dict.get(phone, 0.0),
                        'lastVisit': c[7].isoformat() if c[7] else None
                    }

        result = list(customers_dict.values())

        return jsonify({
            'success': True,
            'customers': result
        }), 200

    except Exception as e:
        print(f"Get customers error: {str(e)}")
        return jsonify({"error": "Failed to fetch customers"}), 400


# ------------------ ADD / CREATE NEW CUSTOMER ------------------
@billing_bp.route("/billing/customers", methods=["POST"])
def add_new_customer():
    """Add a new customer manually or save details"""
    try:
        data = request.get_json() or {}
        phone = data.get('phone') or data.get('customerPhone')
        name = data.get('name') or data.get('customerName') or 'Walk-in Customer'

        if not phone:
            return jsonify({"error": "Phone number is required"}), 400

        rewards = CustomerRewards.query.filter_by(phone=phone).first()
        if not rewards:
            rewards = CustomerRewards()
            rewards.phone = phone
            rewards.total_points_earned = 0.0
            rewards.total_points_redeemed = 0.0
            rewards.current_balance = 0.0
            rewards.total_spend = 0.0
            rewards.bill_count = 0
            db.session.add(rewards)

        # Always update profile fields
        if name and name != 'Walk-in Customer':
            rewards.name = name
        rewards.first_name = data.get('firstName') or data.get('first_name') or rewards.first_name or ''
        rewards.last_name = data.get('lastName') or data.get('last_name') or rewards.last_name or ''
        rewards.email = data.get('email') or data.get('customerEmail') or rewards.email or ''
        rewards.gst = data.get('gst') or data.get('customerGST') or rewards.gst or ''
        rewards.address = data.get('address') or data.get('customerAddress') or rewards.address or ''
        rewards.customer_type = data.get('type') or data.get('customerType') or rewards.customer_type or 'regular'
        rewards.member_id = data.get('memberId') or data.get('member_id') or rewards.member_id or ''
        rewards.is_classic_customer = bool(data.get('isClassicCustomer', rewards.is_classic_customer or False))
        rewards.is_supplier = bool(data.get('isSupplier', rewards.is_supplier or False))
        rewards.supplier_igst = float(data.get('supplierIGST', rewards.supplier_igst or 0))

        # Parse optional date fields
        from datetime import date
        def parse_date(val):
            if not val: return None
            try:
                return date.fromisoformat(str(val))
            except Exception:
                return None

        rewards.date_of_birth = parse_date(data.get('dateOfBirth')) or rewards.date_of_birth
        rewards.wedding_anniversary = parse_date(data.get('weddingAnniversary')) or rewards.wedding_anniversary
        rewards.celebration_date = parse_date(data.get('celebrationDate')) or rewards.celebration_date

        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Customer saved successfully",
            "customer": rewards.to_dict()
        }), 201

    except Exception as e:
        db.session.rollback()
        print(f"Add customer error: {str(e)}")
        return jsonify({"error": str(e)}), 400


# ------------------ UPDATE EXISTING CUSTOMER ------------------
@billing_bp.route("/billing/customers/<path:phone>", methods=["PUT"])
def update_customer(phone):
    """Update an existing customer profile by phone number"""
    try:
        data = request.get_json() or {}
        phone = str(phone).strip()
        rewards = CustomerRewards.query.filter_by(phone=phone).first()

        new_phone = str(data.get('phone') or data.get('customerPhone') or phone).strip()
        name = data.get('name') or data.get('customerName') or (f"{data.get('firstName', '')} {data.get('lastName', '')}".strip()) or 'Walk-in Customer'

        if not rewards:
            rewards = CustomerRewards()
            rewards.phone = new_phone
            rewards.total_points_earned = 0.0
            rewards.total_points_redeemed = 0.0
            rewards.current_balance = 0.0
            rewards.total_spend = 0.0
            rewards.bill_count = 0
            db.session.add(rewards)
        elif new_phone != phone:
            existing = CustomerRewards.query.filter_by(phone=new_phone).first()
            if existing and existing != rewards:
                return jsonify({"error": f"Customer with phone number {new_phone} already exists"}), 400
            rewards.phone = new_phone

        if name:
            rewards.name = name
        rewards.first_name = data.get('firstName') if 'firstName' in data else (data.get('first_name') if 'first_name' in data else rewards.first_name)
        rewards.last_name = data.get('lastName') if 'lastName' in data else (data.get('last_name') if 'last_name' in data else rewards.last_name)
        rewards.email = data.get('email') if 'email' in data else (data.get('customerEmail') if 'customerEmail' in data else rewards.email)
        rewards.gst = data.get('gst') if 'gst' in data else (data.get('customerGST') if 'customerGST' in data else rewards.gst)
        rewards.address = data.get('address') if 'address' in data else (data.get('customerAddress') if 'customerAddress' in data else rewards.address)
        rewards.customer_type = data.get('type') if 'type' in data else (data.get('customerType') if 'customerType' in data else rewards.customer_type)
        rewards.member_id = data.get('memberId') if 'memberId' in data else (data.get('member_id') if 'member_id' in data else rewards.member_id)
        if 'isClassicCustomer' in data:
            rewards.is_classic_customer = bool(data['isClassicCustomer'])
        if 'isSupplier' in data:
            rewards.is_supplier = bool(data['isSupplier'])
        if 'supplierIGST' in data:
            rewards.supplier_igst = float(data['supplierIGST'] or 0)

        if 'rewardPoints' in data:
            val = float(data['rewardPoints'] or 0)
            rewards.current_balance = val
            if val == 0:
                rewards.total_points_earned = 0.0
                rewards.total_points_redeemed = 0.0

        from datetime import date
        def parse_date(val):
            if not val: return None
            try:
                if 'T' in str(val):
                    val = str(val).split('T')[0]
                return date.fromisoformat(str(val))
            except Exception:
                return None

        if 'dateOfBirth' in data:
            rewards.date_of_birth = parse_date(data.get('dateOfBirth'))
        if 'weddingAnniversary' in data:
            rewards.wedding_anniversary = parse_date(data.get('weddingAnniversary'))
        if 'celebrationDate' in data:
            rewards.celebration_date = parse_date(data.get('celebrationDate'))

        # Also update matching bills with new customer details if updated
        matching_bills = Bill.query.filter_by(customer_phone=phone).all()
        for b in matching_bills:
            if new_phone != phone:
                b.customer_phone = new_phone
            if name:
                b.customer_name = name
            if rewards.email:
                b.customer_email = rewards.email
            if rewards.gst:
                b.customer_gst = rewards.gst
            if rewards.address:
                b.customer_address = rewards.address

        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Customer updated successfully",
            "customer": rewards.to_dict()
        }), 200

    except Exception as e:
        db.session.rollback()
        print(f"Update customer error: {str(e)}")
        return jsonify({"error": str(e)}), 400


# ------------------ DELETE CUSTOMER ------------------
@billing_bp.route("/billing/customers/<path:phone>", methods=["DELETE"])
def delete_customer(phone):
    """Delete a customer record by phone number"""
    try:
        phone = str(phone).strip()
        rewards = CustomerRewards.query.filter_by(phone=phone).first()
        if rewards:
            db.session.delete(rewards)
            
        # Also clear customer_phone from associated bills so the customer doesn't reappear in consolidated list
        matching_bills = Bill.query.filter_by(customer_phone=phone).all()
        for b in matching_bills:
            b.customer_phone = ''

        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Customer deleted successfully"
        }), 200

    except Exception as e:
        db.session.rollback()
        print(f"Delete customer error: {str(e)}")
        return jsonify({"error": str(e)}), 400


# ------------------ IMPORT CUSTOMERS (Excel / CSV) ------------------
@billing_bp.route("/billing/customers/import", methods=["POST"])
def import_customers():
    """Import customer records from Excel (.xlsx/.xls) or CSV file"""
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file uploaded. Please attach a file under key 'file'"}), 400

        file = request.files['file']
        if not file or not file.filename:
            return jsonify({"error": "No file selected"}), 400

        filename = file.filename.lower()
        rows_data = []

        # Read Excel (.xlsx, .xls) or CSV
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file, data_only=True)
                sheet = wb.active
                raw_rows = list(sheet.iter_rows(values_only=True))
                if not raw_rows or len(raw_rows) < 2:
                    return jsonify({"error": "Excel file is empty or missing data rows"}), 400
                
                headers = [str(h).strip() if h is not None else '' for h in raw_rows[0]]
                for idx, row in enumerate(raw_rows[1:], start=2):
                    row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                    rows_data.append((idx, row_dict))
            except Exception as ex_err:
                return jsonify({"error": f"Failed to read Excel file: {str(ex_err)}"}), 400
        elif filename.endswith('.csv'):
            import csv, io
            stream = io.StringIO(file.stream.read().decode("utf-8", errors="ignore"), newline=None)
            reader = csv.DictReader(stream)
            for idx, row_dict in enumerate(reader, start=2):
                rows_data.append((idx, row_dict))
        else:
            return jsonify({"error": "Unsupported file format. Please upload an Excel (.xlsx/.xls) or CSV file"}), 400

        # Helper normalizer for keys
        def norm_key(k):
            return str(k or '').strip().lower().replace(' ', '_').replace('-', '_')

        def clean_phone(val):
            if val is None: return ''
            if isinstance(val, float): val = int(val)
            s = str(val).strip()
            if s.endswith('.0'): s = s[:-2]
            return s

        def parse_date_val(val):
            if not val: return None
            from datetime import date, datetime
            if isinstance(val, (datetime, date)):
                return val if isinstance(val, date) else val.date()
            val_str = str(val).strip()
            if not val_str or val_str.lower() in ('none', 'null', 'n/a', '-', ''):
                return None
            for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d', '%d.%m.%Y', '%m/%d/%Y'):
                try:
                    return datetime.strptime(val_str, fmt).date()
                except ValueError:
                    pass
            return 'INVALID_DATE'

        def parse_bool_val(val):
            if isinstance(val, bool): return val
            if isinstance(val, (int, float)): return bool(val)
            s = str(val or '').strip().lower()
            return s in ('true', '1', 'yes', 'y', 'classic', 'supplier')

        def parse_float_val(val, default=0.0):
            if val is None or val == '': return default
            try: return float(val)
            except ValueError: return default

        imported_count = 0
        updated_count = 0
        errors = []

        for row_num, row in rows_data:
            # Map normalized headers
            norm_map = {norm_key(k): v for k, v in row.items() if k}

            raw_phone = norm_map.get('mobile_number') or norm_map.get('mobilenumber') or norm_map.get('phone') or norm_map.get('mobile') or norm_map.get('phone_number')
            phone = clean_phone(raw_phone)

            # Validation: Mobile Number is required
            if not phone or len(phone) < 5:
                errors.append(f"Row {row_num}: Mobile_Number is missing or invalid ('{raw_phone or ''}')")
                continue

            first_name = str(norm_map.get('first_name') or norm_map.get('firstname') or '').strip()
            last_name = str(norm_map.get('last_name') or norm_map.get('lastname') or '').strip()
            email = str(norm_map.get('email_id') or norm_map.get('emailid') or norm_map.get('email') or '').strip()
            address = str(norm_map.get('address') or '').strip()
            member_id = str(norm_map.get('member_id') or norm_map.get('memberid') or '').strip()
            gst = str(norm_map.get('gst') or norm_map.get('gst_number') or '').strip()

            dob = parse_date_val(norm_map.get('date_of_birth') or norm_map.get('dateofbirth') or norm_map.get('dob'))
            anniversary = parse_date_val(norm_map.get('wedding_anniversary') or norm_map.get('anniversary'))
            celebration = parse_date_val(norm_map.get('celebration_date') or norm_map.get('celebration'))

            if dob == 'INVALID_DATE':
                errors.append(f"Row {row_num}: Invalid Date_of_Birth format")
                dob = None
            if anniversary == 'INVALID_DATE':
                errors.append(f"Row {row_num}: Invalid Wedding_Anniversary format")
                anniversary = None
            if celebration == 'INVALID_DATE':
                errors.append(f"Row {row_num}: Invalid Celebration_Date format")
                celebration = None

            is_classic = parse_bool_val(norm_map.get('isclassic_customer') or norm_map.get('is_classic_customer') or norm_map.get('classic_customer'))
            is_supplier = parse_bool_val(norm_map.get('issupplier') or norm_map.get('is_supplier') or norm_map.get('supplier'))

            rewards_pt = parse_float_val(norm_map.get('rewards_point') or norm_map.get('rewards_points') or norm_map.get('reward_points') or norm_map.get('points'))
            supplier_igst = parse_float_val(norm_map.get('issupplier_igst') or norm_map.get('supplier_igst') or norm_map.get('supplierigst'))

            # Check if customer exists by phone
            rewards = CustomerRewards.query.filter_by(phone=phone).first()
            is_new = False
            if not rewards:
                rewards = CustomerRewards()
                rewards.phone = phone
                rewards.total_points_earned = 0.0
                rewards.total_points_redeemed = 0.0
                rewards.current_balance = rewards_pt
                rewards.total_spend = 0.0
                rewards.bill_count = 0
                db.session.add(rewards)
                is_new = True
            else:
                if rewards_pt > 0:
                    rewards.current_balance = rewards_pt

            # Update profile fields
            full_name = f"{first_name} {last_name}".strip()
            if full_name:
                rewards.name = full_name
            elif not rewards.name:
                rewards.name = 'Walk-in Customer'

            if first_name: rewards.first_name = first_name
            if last_name:  rewards.last_name = last_name
            if email:      rewards.email = email
            if address:    rewards.address = address
            if gst:        rewards.gst = gst
            if member_id:  rewards.member_id = member_id
            if dob:        rewards.date_of_birth = dob
            if anniversary: rewards.wedding_anniversary = anniversary
            if celebration: rewards.celebration_date = celebration

            rewards.is_classic_customer = is_classic
            rewards.is_supplier = is_supplier
            rewards.supplier_igst = supplier_igst
            rewards.customer_type = 'classic' if is_classic else (rewards.customer_type or 'regular')

            if is_new:
                imported_count += 1
            else:
                updated_count += 1

        db.session.commit()

        return jsonify({
            "success": True,
            "message": f"Successfully processed {imported_count + updated_count} records",
            "importedCount": imported_count,
            "updatedCount": updated_count,
            "totalProcessed": len(rows_data),
            "errors": errors
        }), 200

    except Exception as e:
        db.session.rollback()
        print(f"Import customers error: {str(e)}")
        print(traceback.format_exc())
        return jsonify({"error": f"Failed to import customers: {str(e)}"}), 500


# ------------------ EXPORT CUSTOMERS (Excel .xlsx) ------------------
@billing_bp.route("/billing/customers/export-excel", methods=["GET"])
def export_customers_excel():
    """Export customer records as a binary Excel (.xlsx) file with specified columns"""
    try:
        import openpyxl, io
        from flask import send_file

        rewards = CustomerRewards.query.all()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customers"

        headers = [
            'First_Name', 'Last_Name', 'Date_of_Birth', 'Mobile_Number', 'Email_ID',
            'Address', 'Rewards_Point', 'Member_ID', 'Wedding_Anniversary',
            'Celebration_Date', 'isClassic_Customer', 'ISSUPPLIER', 'ISSUPPLIER_IGST'
        ]
        ws.append(headers)

        for r in rewards:
            f_name = r.first_name or ''
            l_name = r.last_name or ''
            if not f_name and r.name:
                parts = r.name.strip().split(' ')
                f_name = parts[0] or ''
                l_name = ' '.join(parts[1:]) or ''

            ws.append([
                f_name,
                l_name,
                r.date_of_birth.isoformat() if r.date_of_birth else '',
                r.phone or '',
                r.email or '',
                r.address or '',
                round(r.current_balance or 0, 2),
                r.member_id or '',
                r.wedding_anniversary.isoformat() if r.wedding_anniversary else '',
                r.celebration_date.isoformat() if r.celebration_date else '',
                'TRUE' if r.is_classic_customer else 'FALSE',
                'TRUE' if r.is_supplier else 'FALSE',
                round(r.supplier_igst or 0, 2)
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="customer_details.xlsx"
        )
    except Exception as e:
        print(f"Export customers error: {str(e)}")
# ------------------ RESET ALL CUSTOMER POINTS ------------------
@billing_bp.route("/billing/customers/reset-all-points", methods=["POST"])
def reset_all_customer_points():
    """Reset reward points for all customers in the database to 0"""
    try:
        updated_count = CustomerRewards.query.update({
            CustomerRewards.current_balance: 0.0,
            CustomerRewards.total_points_earned: 0.0,
            CustomerRewards.total_points_redeemed: 0.0
        })
        db.session.commit()
        return jsonify({
            "success": True,
            "message": f"Successfully reset reward points to 0 for {updated_count} customers."
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Failed to reset customer points: {str(e)}"}), 500


# ------------------ RESET SINGLE CUSTOMER POINTS ------------------
@billing_bp.route("/billing/customers/<path:phone>/reset-points", methods=["POST"])
def reset_single_customer_points(phone):
    """Reset reward points for a specific customer in the database to 0"""
    try:
        phone = str(phone).strip()
        clean_p = ''.join(filter(str.isdigit, phone))
        
        rewards = CustomerRewards.query.filter_by(phone=phone).first()
        if not rewards and phone:
            rewards = CustomerRewards.query.filter(
                (CustomerRewards.phone == phone) |
                (CustomerRewards.member_id == phone)
            ).first()
        if not rewards and clean_p:
            rewards = CustomerRewards.query.filter(
                (CustomerRewards.phone == clean_p) |
                (CustomerRewards.member_id == clean_p)
            ).first()

        if rewards:
            rewards.current_balance = 0.0
            rewards.total_points_earned = 0.0
            rewards.total_points_redeemed = 0.0
            db.session.commit()
            return jsonify({
                "success": True,
                "message": f"Successfully reset reward points to 0 for customer ({rewards.phone or phone})."
            }), 200
        else:
            return jsonify({"success": True, "message": "Customer record not found in DB; local points reset to 0."}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Failed to reset customer points: {str(e)}"}), 500





# ------------------ CREATE NEW BILL ------------------
@billing_bp.route("/billing/bills", methods=["POST"])
def create_bill():
    """Create a new bill with items and payment"""
    try:
        data = request.get_json()
        
        # Validate required fields
        if not data.get('items'):
            return jsonify({"error": "No items in bill"}), 400
            
        if len(data['items']) == 0:
            return jsonify({"error": "Bill must have at least one item"}), 400
        
        # Create new bill instance with unique number
        bill = Bill()
        req_bill_no = str(data.get('billNumber', '')).strip()
        if req_bill_no:
            if re.match(r'^[A-Z]+\d+[NR]$', req_bill_no.upper()):
                bill.bill_number = req_bill_no.upper()
            elif '/' in req_bill_no:
                bill.bill_number = req_bill_no.split('/')[-1].upper()
            else:
                fy_letter = get_fy_letter()
                clean_no = req_bill_no.rstrip('N').rstrip('n').rstrip('R').rstrip('r')
                clean_digits = re.sub(r'\D', '', clean_no)
                suffix = 'R' if req_bill_no.upper().endswith('R') else 'N'
                if clean_digits:
                    num = int(clean_digits)
                    bill.bill_number = f"{fy_letter}{num:04d}{suffix}"
                else:
                    bill.bill_number = req_bill_no.upper()
        else:
            bill.bill_number = generate_unique_bill_number('N')
        
        # Customer Information
        bill.customer_name = data.get('customerName', 'Walk-in Customer')
        bill.customer_phone = data.get('customerPhone', '')
        bill.customer_email = data.get('customerEmail', '')
        bill.customer_gst = data.get('customerGST', '')
        bill.customer_address = data.get('customerAddress')
        bill.contact = data.get('contact')
        bill.customer_type = data.get('customerType', 'regular')
        
        # Vehicle Information
        bill.vehicle_name = data.get('vehicleName', '')
        bill.vehicle_number = data.get('vehicleNumber', '')
        
        # Company Information - Fetch and store snapshot
        company_id = data.get('companyId')
        if company_id:
            company = Company.query.get(company_id)
            if company:
                bill.company_id = company.id
                bill.company_name = company.name
                bill.company_address = company.address
                bill.company_phone = company.phone
                bill.company_email = company.email
                bill.company_gst = company.gst_number
                bill.company_alternate_phone = company.alternate_phone
                bill.company_bank_name = company.bank_name
                bill.company_bank_account = company.bank_account_number
                bill.company_bank_ifsc = company.bank_ifsc
                bill.company_bank_branch = company.bank_branch
                bill.company_upi_id = company.upi_id
                # Store logo path if exists
                if hasattr(company, 'logo_path') and company.logo_path:
                    bill.company_logo = company.logo_path
        
        # Created By (User information) - Hide discount details from employee
        bill.created_by = data.get('createdBy', None)
        bill.created_by_name = data.get('createdByName', 'System')
        
        # Discount and tax settings (employee should not see discount details)
        # These will be applied but not shown to employee
        bill.discount = float(data.get('discount', 0))
        bill.discount_type = data.get('discountType', 'amount')  # 'amount' or 'percentage'
        bill.tax = float(data.get('tax', 0))
        bill.tax_type = data.get('taxType', 'percentage')
        
        # Payment information
        bill.subtotal = float(data.get('subtotal', 0))
        bill.paid_amount = float(data.get('paidAmount', 0))
        bill.payment_method = data.get('paymentMethod', 'cash')
        
        # Payment details snapshot
        bill.cash_received = float(data.get('cashReceived', 0))
        bill.payment_card_number = data.get('cardNumber', '')
        bill.payment_card_holder = data.get('cardHolderName', '')
        bill.payment_upi_id = data.get('upiId', '')
        bill.payment_transaction_id = data.get('transactionId', '')
        bill.payment_bank_name = data.get('bankName', '')
        bill.payment_cheque_number = data.get('chequeNumber', '')
        bill.payment_online_phone = data.get('onlinePhone', '')
        bill.payment_online_ref = data.get('onlineRef', '')
        if hasattr(bill, 'sales_return_amount'):
            bill.sales_return_amount = float(data.get('salesReturnAmount', 0))
        
        # Add items and update stock
        items_added = []
        for item_data in data.get('items', []):
            product = Product.query.get(item_data['productId'])
            
            if not product:
                db.session.rollback()
                return jsonify({"error": f"Product with ID {item_data['productId']} not found"}), 404
            
            quantity = int(item_data['quantity'])
            if quantity <= 0:
                db.session.rollback()
                return jsonify({"error": f"Invalid quantity for {product.name}"}), 400
                
            if product.quantity < quantity:
                db.session.rollback()
                return jsonify({"error": f"Insufficient stock for {product.name}. Available: {product.quantity}"}), 400
            
            # Use price from frontend if provided (for overrides), otherwise use product price
            line_price = item_data.get('price')
            if line_price is None:
                line_price = product.net_price or product.sell_price
            
            item_total = float(line_price) * quantity
            
            # Determine which profit to use based on isClassic flag
            is_classic_bill = data.get('isClassic', False)
            
            if is_classic_bill:
                # Classic Customer Profit = Classic Customer Price - Purchase Rate (buy_price)
                profit_base = float(product.classic_customer or 0)
            else:
                # Normal Profit = MRP - Purchase Rate (buy_price)
                profit_base = float(product.mrp or 0)
            
            unit_profit = profit_base - (product.buy_price or 0)
            item_profit = unit_profit * quantity
            
            # Create bill item with status (defaults to 'pending' from model)
            bill_item = BillItem()
            bill_item.product_id = product.id
            bill_item.product_code = product.product_code
            bill_item.product_name = product.name
            bill_item.product_model = product.model or ''
            bill_item.product_type = product.type or ''
            bill_item.tax = float(item_data.get('tax') or product.tax or 5.0)
            bill_item.sell_price = line_price
            bill_item.buy_price = product.buy_price or 0
            bill_item.quantity = quantity
            bill_item.total = item_total
            bill_item.profit = item_profit
            
            # Update product quantity
            product.quantity -= quantity
            
            bill.items.append(bill_item)
            items_added.append({
                'name': product.name,
                'quantity': quantity,
                'total': item_total,
                'status': 'pending'
            })
        
        # Calculate all totals (including discount and tax) using inclusive tax method
        bill.calculate_totals(is_tax_inclusive=True)
        if data.get('total') is not None:
            bill.total = float(data['total'])
        if hasattr(bill, 'sales_return_amount'):
            req_sr = float(data.get('salesReturnAmount', 0))
            bill.sales_return_amount = min(req_sr, bill.total)
        
        # Update or Create CustomerRewards
        cust_phone = bill.customer_phone or data.get('contact') or data.get('phone') or data.get('customerPhone')
        member_id = data.get('memberId') or data.get('member_id')
        rewards = None
        if cust_phone:
            rewards = CustomerRewards.query.filter_by(phone=cust_phone).first()
        if not rewards and member_id:
            rewards = CustomerRewards.query.filter_by(member_id=member_id).first()

        points_earned = float(data.get('rewardPointsEarned', 0))
        points_redeemed = float(data.get('rewardPointsRedeemed', 0))

        if rewards or cust_phone:
            if not rewards:
                rewards = CustomerRewards()
                rewards.phone = cust_phone
                rewards.name = bill.customer_name
                rewards.member_id = member_id or ''
                rewards.total_points_earned = points_earned
                rewards.total_points_redeemed = points_redeemed
                rewards.current_balance = max(0.0, points_earned - points_redeemed)
                rewards.total_spend = bill.total
                rewards.bill_count = 1
                db.session.add(rewards)
            else:
                rewards.total_points_earned += points_earned
                rewards.total_points_redeemed += points_redeemed
                rewards.current_balance = max(0.0, rewards.total_points_earned - rewards.total_points_redeemed)
                rewards.total_spend += bill.total
                rewards.bill_count += 1
                if not rewards.name or rewards.name == 'Walk-in Customer':
                    rewards.name = bill.customer_name

        # Save to database
        db.session.add(bill)
        db.session.commit()
        
        # Create payment record if amount paid
        if bill.paid_amount > 0:
            payment = Payment(
                bill_id=bill.id,
                payment_id=f"PAY-{bill.bill_number}",
                amount=bill.paid_amount,
                method=bill.payment_method,
                status='completed' if bill.paid_amount >= bill.total else 'partial'
            )
            db.session.add(payment)
            db.session.commit()
        
        # Return response - hide discount details from employee
        return jsonify({
            'success': True,
            'message': 'Bill created successfully',
            'billNumber': bill.bill_number.split('/', 1)[-1] if '/' in bill.bill_number else bill.bill_number,
            'billId': bill.id,
            'total': round(bill.total, 2),
            'changeAmount': round(bill.change_amount, 2),
            'items': items_added
        }), 201
        
    except Exception as e:
        db.session.rollback()
        print(f"Create bill error: {str(e)}")
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 400


# ------------------ UPDATE CUSTOMER INFORMATION ------------------
@billing_bp.route("/billing/customer/<string:phone_number>", methods=["PUT"])
def update_customer_info(phone_number):
    """Update customer information for all existing records"""
    try:
        data = request.get_json()
        
        if not phone_number:
            return jsonify({"error": "Phone number is required"}), 400
        
        # Find all bills with this phone number and update customer info
        existing_bills = Bill.query.filter_by(customer_phone=phone_number).all()
        
        if not existing_bills:
            return jsonify({"error": "Customer not found"}), 404
        
        # Update all records with new information
        for bill in existing_bills:
            if data.get('name'):
                bill.customer_name = data.get('name')
            if data.get('email'):
                bill.customer_email = data.get('email')
            if data.get('gst'):
                bill.customer_gst = data.get('gst')
            if data.get('address'):
                bill.customer_address = data.get('address')
            if data.get('type'):
                bill.customer_type = data.get('type')
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Customer information updated successfully',
            'customer': {
                'name': existing_bills[0].customer_name,
                'phone': existing_bills[0].customer_phone,
                'email': existing_bills[0].customer_email,
                'gst': existing_bills[0].customer_gst,
                'address': existing_bills[0].customer_address,
                'type': existing_bills[0].customer_type
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Update customer error: {str(e)}")
        return jsonify({"error": "Failed to update customer information"}), 400


# ------------------ GET BILLS WITH PENDING ITEMS ------------------
@billing_bp.route("/billing/bills/pending-items", methods=["GET"])
def get_bills_with_pending_items():
    """Get all bills that have pending items"""
    try:
        # Find all bills that have at least one pending item
        bills = Bill.query.join(BillItem).filter(
            BillItem.item_status == 'pending'
        ).distinct(Bill.id).order_by(Bill.created_at.desc()).all()
        
        result = []
        for bill in bills:
            # Count pending items for this bill
            pending_count = BillItem.query.filter_by(
                bill_id=bill.id, 
                item_status='pending'
            ).count()
            
            result.append({
                'id': bill.id,
                'billNumber': bill.bill_number.split('/', 1)[-1] if (bill.bill_number and '/' in bill.bill_number) else bill.bill_number,
                'customerName': bill.customer_name,
                'customerPhone': bill.customer_phone,
                'customerType': bill.customer_type,
                'vehicleName': bill.vehicle_name,
                'vehicleNumber': bill.vehicle_number,
                'companyName': bill.company_name,
                'total': round(bill.total, 2),
                'paidAmount': round(bill.paid_amount, 2),
                'pendingItems': pending_count,
                'createdAt': bill.created_at.isoformat() if bill.created_at else None,
                'createdBy': bill.created_by,
                'createdByName': bill.created_by_name
            })
        
        return jsonify({
            'success': True,
            'bills': result
        }), 200
        
    except Exception as e:
        print(f"Get pending bills error: {str(e)}")
        print(traceback.format_exc())
        return jsonify({"error": "Failed to fetch pending bills"}), 400


# ------------------ GET PENDING ITEMS FOR A BILL ------------------
@billing_bp.route("/billing/bills/<int:bill_id>/items/pending", methods=["GET"])
def get_pending_bill_items(bill_id):
    """Get all pending items for a specific bill"""
    try:
        bill = Bill.query.get_or_404(bill_id)
        
        pending_items = BillItem.query.filter_by(
            bill_id=bill_id,
            item_status='pending'
        ).all()
        
        items = [item.to_dict() for item in pending_items]
        
        return jsonify({
            'success': True,
            'bill_id': bill_id,
            'bill_number': bill.bill_number,
            'customer_type': bill.customer_type,
            'customer_name': bill.customer_name,
            'vehicle_name': bill.vehicle_name,
            'vehicle_number': bill.vehicle_number,
            'company_name': bill.company_name,
            'items': items
        }), 200
        
    except Exception as e:
        print(f"Get pending items error: {str(e)}")
        return jsonify({"error": "Failed to fetch pending items"}), 400


# ------------------ COMPLETE A BILL ITEM ------------------
@billing_bp.route("/billing/bills/<int:bill_id>/items/<int:item_id>/complete", methods=["POST"])
def complete_bill_item(bill_id, item_id):
    """Mark a bill item as completed (inventory already updated during bill creation)"""
    try:
        bill = Bill.query.get_or_404(bill_id)
        item = BillItem.query.get_or_404(item_id)
        
        if item.bill_id != bill.id:
            return jsonify({"error": "Item does not belong to this bill"}), 400
        
        if item.item_status != 'pending':
            return jsonify({"error": "Item is already completed"}), 400
        
        # Update item status to completed
        item.item_status = 'completed'
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Item marked as completed successfully',
            'item': {
                'id': item.id,
                'status': item.item_status
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Complete item error: {str(e)}")
        return jsonify({"error": str(e)}), 400


# ------------------ COMPLETE ALL ITEMS IN A BILL ------------------
@billing_bp.route("/billing/bills/<int:bill_id>/complete-all", methods=["POST"])
def complete_all_bill_items(bill_id):
    """Mark all pending items in a bill as completed"""
    try:
        bill = Bill.query.get_or_404(bill_id)
        
        # Get all pending items
        pending_items = BillItem.query.filter_by(
            bill_id=bill_id,
            item_status='pending'
        ).all()
        
        if not pending_items:
            return jsonify({"error": "No pending items found in this bill"}), 400
        
        completed_count = 0
        for item in pending_items:
            item.item_status = 'completed'
            completed_count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Successfully completed {completed_count} items',
            'completedCount': completed_count
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Complete all items error: {str(e)}")
        return jsonify({"error": str(e)}), 400


# ------------------ GET CUSTOMER BILLS (unified search + date range) ------------------
@billing_bp.route("/billing/customer-bills", methods=["GET"])
def get_customer_bills_by_phone_date():
    """Search bills by bill_number, phone, or customer name; filter by date range"""
    try:
        search = request.args.get('search', '').strip()
        phone = request.args.get('phone', '').strip()
        from_date_str = request.args.get('from_date', '').strip()
        to_date_str = request.args.get('to_date', '').strip()
        date_str = request.args.get('date', '').strip()

        query = Bill.query
        
        include_returns = request.args.get('include_returns', 'false').strip().lower() == 'true'
        if not include_returns:
            query = query.filter(~or_(Bill.bill_number.endswith('R'), Bill.bill_number.endswith('r')))

        term = search or phone
        if term:
            query = query.filter(
                or_(
                    Bill.customer_phone == term,
                    Bill.bill_number.ilike(f'%{term}%'),
                    Bill.customer_name.ilike(f'%{term}%')
                )
            )

        if from_date_str:
            try:
                fd = datetime.strptime(from_date_str, '%Y-%m-%d')
                query = query.filter(Bill.created_at >= fd)
            except ValueError:
                return jsonify({"error": "Invalid from_date. Use YYYY-MM-DD"}), 400

        if to_date_str:
            try:
                td = datetime.strptime(to_date_str, '%Y-%m-%d')
                td = td.replace(hour=23, minute=59, second=59)
                query = query.filter(Bill.created_at <= td)
            except ValueError:
                return jsonify({"error": "Invalid to_date. Use YYYY-MM-DD"}), 400

        if date_str and not from_date_str and not to_date_str:
            try:
                filter_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                query = query.filter(func.date(Bill.created_at) == filter_date)
            except ValueError:
                return jsonify({"error": "Invalid date. Use YYYY-MM-DD"}), 400

        bills_data = query.order_by(Bill.created_at.asc()).all()

        if not bills_data:
            return jsonify({"bills": [], "customers": {}, "totalBills": 0, "message": "No bills found"}), 200

        def build_bill_dict(bill):
            items = []
            for item in bill.items:
                items.append({
                    "id": item.id,
                    "productName": item.product_name or "",
                    "productCode": item.product_code or "",
                    "description": item.description or "",
                    "quantity": item.quantity or 0,
                    "unit": item.unit or "PCS",
                    "mrp": round(item.mrp or 0, 2),
                    "sellingPrice": round(item.selling_price or 0, 2),
                    "discount": round(item.discount or 0, 2),
                    "tax": round(item.tax or 0, 2),
                    "total": round(item.total or 0, 2),
                    "status": item.item_status or "completed"
                })
            return {
                "id": bill.id,
                "billNumber": bill.bill_number.split('/', 1)[-1] if (bill.bill_number and '/' in bill.bill_number) else bill.bill_number,
                "billDate": bill.created_at.strftime('%d-%m-%Y') if bill.created_at else "",
                "billTime": bill.created_at.strftime('%I:%M %p') if bill.created_at else "",
                "billDateRaw": bill.created_at.isoformat() if bill.created_at else "",
                "customerName": bill.customer_name or "",
                "customerPhone": bill.customer_phone or "",
                "customerEmail": bill.customer_email or "",
                "customerGST": bill.customer_gst or "",
                "customerAddress": bill.customer_address or "",
                "customerType": bill.customer_type or "regular",
                "subtotal": round(bill.subtotal or 0, 2),
                "discount": round(bill.discount or 0, 2),
                "tax": round(bill.tax or 0, 2),
                "total": round(bill.total or 0, 2),
                "paidAmount": round(bill.paid_amount or 0, 2),
                "paymentMethod": bill.payment_method or "",
                "paymentStatus": bill.payment_status or "",
                "createdBy": bill.created_by_name or bill.created_by or "",
                "items": items
            }

        result = [build_bill_dict(b) for b in bills_data]

        customers_map = {}
        for b in result:
            key = b["customerPhone"] or b["customerName"]
            if key not in customers_map:
                customers_map[key] = {
                    "name": b["customerName"],
                    "phone": b["customerPhone"],
                    "email": b["customerEmail"],
                    "gst": b["customerGST"],
                    "address": b["customerAddress"],
                    "type": b["customerType"],
                }

        return jsonify({
            "bills": result,
            "customers": customers_map,
            "totalBills": len(result)
        }), 200

    except Exception as e:
        print(f"Get customer bills error: {str(e)}")
        print(traceback.format_exc())
        return jsonify({"error": "Failed to fetch customer bills"}), 500


# ------------------ GET ALL BILLS (with pagination) ------------------
@billing_bp.route("/billing/bills", methods=["GET"])
def get_all_bills():
    """Get all bills with pagination and filters"""
    try:
        # Pagination parameters
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        
        # Filter parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        customer = request.args.get('customer')
        customer_type = request.args.get('customer_type')
        vehicle_number = request.args.get('vehicle_number')
        payment_method = request.args.get('payment_method')
        payment_status = request.args.get('payment_status')
        company_id = request.args.get('company_id', type=int)
        
        # Build query
        query = Bill.query
        
        # Exclude Sales Return bills from normal bill lists unless type='R' or include_returns='true'
        bill_type_filter = request.args.get('type', 'N').strip().upper()
        include_returns = request.args.get('include_returns', 'false').strip().lower() == 'true'

        if not include_returns:
            if bill_type_filter == 'R':
                query = query.filter(or_(Bill.bill_number.endswith('R'), Bill.bill_number.endswith('r')))
            else:
                query = query.filter(~or_(Bill.bill_number.endswith('R'), Bill.bill_number.endswith('r')))
        
        if start_date:
            query = query.filter(Bill.created_at >= datetime.fromisoformat(start_date))
        if end_date:
            query = query.filter(Bill.created_at <= datetime.fromisoformat(end_date))
        if customer:
            query = query.filter(Bill.customer_name.ilike(f'%{customer}%'))
        if customer_type:
            query = query.filter(Bill.customer_type == customer_type)
        if vehicle_number:
            query = query.filter(Bill.vehicle_number.ilike(f'%{vehicle_number}%'))
        if payment_method:
            query = query.filter(Bill.payment_method == payment_method)
        if payment_status:
            query = query.filter(Bill.payment_status == payment_status)
        if company_id:
            query = query.filter(Bill.company_id == company_id)
        
        # Order by most recent first
        query = query.order_by(Bill.created_at.desc())
        
        # Paginate
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        
        # Format response
        bills = []
        for bill in pagination.items:
            # Count pending items
            pending_count = BillItem.query.filter_by(
                bill_id=bill.id, 
                item_status='pending'
            ).count()
            
            bills.append({
                'id': bill.id,
                'billNumber': bill.bill_number.split('/', 1)[-1] if (bill.bill_number and '/' in bill.bill_number) else bill.bill_number,
                'customerName': bill.customer_name,
                'customerPhone': bill.customer_phone,
                'customerType': bill.customer_type,
                'customerEmail': bill.customer_email,
                'customerGST': bill.customer_gst,
                'vehicleName': bill.vehicle_name,
                'vehicleNumber': bill.vehicle_number,
                'companyName': bill.company_name,
                'companyGST': bill.company_gst,
                'contact': bill.contact,
                'subtotal': round(bill.subtotal or 0, 2),
                'discount': round(bill.discount or 0, 2),
                'tax': round(bill.tax or 0, 2),
                'total': round(bill.total or 0, 2),
                'profit': round(bill.profit or 0, 2),
                'paidAmount': round(bill.paid_amount or 0, 2),
                'paymentMethod': bill.payment_method,
                'paymentStatus': bill.payment_status,
                'itemCount': len(bill.items),
                'totalQuantity': sum(int(item.quantity or 1) for item in bill.items) if bill.items else max(1, len(bill.items)),
                'items': [item.to_dict() for item in bill.items],
                'pendingItems': pending_count,
                'createdAt': bill.created_at.isoformat() if bill.created_at else None,
                'createdBy': bill.created_by,
                'createdByName': bill.created_by_name
            })
        
        return jsonify({
            'bills': bills,
            'total': pagination.total,
            'pages': pagination.pages,
            'currentPage': page,
            'perPage': per_page
        }), 200
        
    except Exception as e:
        print(f"Get bills error: {str(e)}")
        print(traceback.format_exc())
        return jsonify({"error": "Failed to fetch bills"}), 400


# ------------------ GET SINGLE BILL BY ID ------------------
@billing_bp.route("/billing/bills/<int:bill_id>", methods=["GET"])
def get_bill_by_id(bill_id):
    """Get detailed bill information by ID"""
    try:
        bill = Bill.query.get_or_404(bill_id)
        
        # Get payment history
        payments = Payment.query.filter_by(bill_id=bill.id).all()
        
        # Get all items with their status
        items = [item.to_dict() for item in bill.items]
        
        bill_dict = bill.to_dict()
        bill_dict['items'] = items
        bill_dict['payments'] = [p.to_dict() for p in payments]
        bill_dict['vehicleName'] = bill.vehicle_name
        bill_dict['vehicleNumber'] = bill.vehicle_number
        bill_dict['createdBy'] = bill.created_by
        bill_dict['createdByName'] = bill.created_by_name
        
        # Add company details to response
        bill_dict['company'] = {
            'id': bill.company_id,
            'name': bill.company_name,
            'address': bill.company_address,
            'city': bill.company_city,
            'phone': bill.company_phone,
            'email': bill.company_email,
            'gst': bill.company_gst,
            'alternatePhone': bill.company_alternate_phone,
            'bankName': bill.company_bank_name,
            'bankAccount': bill.company_bank_account,
            'bankIfsc': bill.company_bank_ifsc,
            'bankBranch': bill.company_bank_branch,
            'upiId': bill.company_upi_id
        }
        
        # Add payment details to response
        bill_dict['paymentDetails'] = {
            'cardNumber': bill.payment_card_number,
            'cardHolder': bill.payment_card_holder,
            'upiId': bill.payment_upi_id,
            'transactionId': bill.payment_transaction_id,
            'bankName': bill.payment_bank_name,
            'chequeNumber': bill.payment_cheque_number,
            'cashReceived': bill.cash_received
        }
        
        return jsonify(bill_dict), 200
        
    except Exception as e:
        print(f"Get bill error: {str(e)}")
        return jsonify({"error": "Bill not found"}), 404


# ------------------ GET BILL BY NUMBER ------------------
@billing_bp.route("/billing/bills/number/<path:bill_number>", methods=["GET"])
def get_bill_by_number(bill_number):
    """Get bill by bill number (supports path with / like 26-27/1N or simple number 1N)"""
    try:
        clean_no = str(bill_number).strip()
        bill = Bill.query.filter_by(bill_number=clean_no).first()
        if not bill:
            fy_prefix = get_current_financial_year()
            raw_seq = clean_no.split('/')[-1].rstrip('N').rstrip('n').rstrip('R').rstrip('r')
            if raw_seq.isdigit():
                num = int(raw_seq)
                possible = [
                    f"{fy_prefix}/{num}N",
                    f"{fy_prefix}/{num}R",
                    f"{fy_prefix}/{raw_seq.zfill(4)}N",
                    f"{num}N",
                    f"{raw_seq.zfill(4)}N",
                    clean_no
                ]
                bill = Bill.query.filter(Bill.bill_number.in_(possible)).first()
        if not bill:
            return jsonify({"error": "Bill not found"}), 404
        
        # Get all items with their status
        items = [item.to_dict() for item in bill.items]
        
        bill_dict = bill.to_dict()
        bill_dict['items'] = items
        bill_dict['vehicleName'] = bill.vehicle_name
        bill_dict['vehicleNumber'] = bill.vehicle_number
        bill_dict['createdBy'] = bill.created_by
        bill_dict['createdByName'] = bill.created_by_name
        
        # Add company details to response
        bill_dict['company'] = {
            'id': bill.company_id,
            'name': bill.company_name,
            'address': bill.company_address,
            'city': bill.company_city,
            'phone': bill.company_phone,
            'email': bill.company_email,
            'gst': bill.company_gst,
            'alternatePhone': bill.company_alternate_phone,
            'bankName': bill.company_bank_name,
            'bankAccount': bill.company_bank_account,
            'bankIfsc': bill.company_bank_ifsc,
            'bankBranch': bill.company_bank_branch,
            'upiId': bill.company_upi_id
        }
        
        # Add payment details to response
        bill_dict['paymentDetails'] = {
            'cardNumber': bill.payment_card_number,
            'cardHolder': bill.payment_card_holder,
            'upiId': bill.payment_upi_id,
            'transactionId': bill.payment_transaction_id,
            'bankName': bill.payment_bank_name,
            'chequeNumber': bill.payment_cheque_number,
            'cashReceived': bill.cash_received
        }
        
        return jsonify(bill_dict), 200
        
    except Exception as e:
        print(f"Get bill by number error: {str(e)}")
        return jsonify({"error": "Bill not found"}), 404


# ------------------ UPDATE BILL PAYMENT ------------------
@billing_bp.route("/billing/bills/<int:bill_id>/payment", methods=["PUT"])
def update_bill_payment(bill_id):
    """Update payment information for a bill"""
    try:
        bill = Bill.query.get_or_404(bill_id)
        data = request.get_json()
        
        # Update payment details
        bill.paid_amount = float(data.get('paidAmount', bill.paid_amount))
        bill.payment_method = data.get('paymentMethod', bill.payment_method)
        
        # Update payment details snapshot
        if 'cashReceived' in data:
            bill.cash_received = float(data.get('cashReceived', 0))
        if 'cardNumber' in data:
            bill.payment_card_number = data.get('cardNumber', '')
        if 'cardHolderName' in data:
            bill.payment_card_holder = data.get('cardHolderName', '')
        if 'upiId' in data:
            bill.payment_upi_id = data.get('upiId', '')
        if 'transactionId' in data:
            bill.payment_transaction_id = data.get('transactionId', '')
        if 'bankName' in data:
            bill.payment_bank_name = data.get('bankName', '')
        if 'chequeNumber' in data:
            bill.payment_cheque_number = data.get('chequeNumber', '')
        
        # Recalculate
        bill.calculate_totals()
        
        # Add payment record
        payment = Payment(
            bill_id=bill.id,
            payment_id=f"PAY-{bill.bill_number}-{datetime.now().strftime('%H%M%S')}",
            amount=data.get('additionalAmount', bill.paid_amount),
            method=bill.payment_method,
            status='completed',
            reference=data.get('reference', ''),
            notes=data.get('notes', '')
        )
        
        db.session.add(payment)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Payment updated successfully',
            'bill': bill.to_dict()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Update payment error: {str(e)}")
        return jsonify({"error": str(e)}), 400


# ------------------ CANCEL/REFUND BILL ------------------
@billing_bp.route("/billing/bills/<int:bill_id>/cancel", methods=["POST"])
def cancel_bill(bill_id):
    """Cancel a bill and restore stock"""
    try:
        bill = Bill.query.get_or_404(bill_id)
        
        # Restore product quantities for items that are not completed
        for item in bill.items:
            if item.item_status != 'completed':
                product = Product.query.get(item.product_id)
                if product:
                    product.quantity += item.quantity
        
        # Update payment status
        for payment in bill.payments:
            payment.status = 'refunded'
        
        # Delete bill (or mark as cancelled)
        db.session.delete(bill)  # Or add a 'cancelled' field to Bill model
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Bill cancelled successfully'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Cancel bill error: {str(e)}")
        return jsonify({"error": str(e)}), 400


# ------------------ DELETE BILL PERMANENTLY ------------------
@billing_bp.route("/billing/bills/<int:bill_id>", methods=["DELETE"])
def delete_bill(bill_id):
    """Permanently delete a bill from the database"""
    try:
        bill = Bill.query.get(bill_id)
        if not bill:
            return jsonify({"error": "Bill not found"}), 404
        
        # Delete related payments if any
        if hasattr(bill, 'payments') and bill.payments:
            for payment in list(bill.payments):
                db.session.delete(payment)
        
        # Delete related bill items
        if hasattr(bill, 'items') and bill.items:
            for item in list(bill.items):
                db.session.delete(item)
                
        db.session.delete(bill)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Bill #{bill.bill_number} permanently deleted from database'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Delete bill error: {str(e)}")
        return jsonify({"error": str(e)}), 500


# ------------------ GET BILLING STATISTICS ------------------
@billing_bp.route("/billing/statistics", methods=["GET"])
def get_billing_statistics():
    """Get billing statistics for dashboard"""
    try:
        # Date range
        today = datetime.now().date()
        start_of_day = datetime(today.year, today.month, today.day, 0, 0, 0)
        end_of_day = datetime(today.year, today.month, today.day, 23, 59, 59)
        
        start_of_week = today - timedelta(days=today.weekday())
        start_of_week = datetime(start_of_week.year, start_of_week.month, start_of_week.day, 0, 0, 0)
        
        start_of_month = datetime(today.year, today.month, 1, 0, 0, 0)
        
        not_return_clause = ~or_(Bill.bill_number.endswith('R'), Bill.bill_number.endswith('r'))

        # Today's stats
        today_stats = db.session.query(
            func.count(Bill.id).label('bill_count'),
            func.sum(Bill.total).label('total_sales'),
            func.avg(Bill.total).label('avg_bill_value')
        ).filter(not_return_clause, Bill.created_at.between(start_of_day, end_of_day)).first()
        
        # Week's stats
        week_stats = db.session.query(
            func.count(Bill.id).label('bill_count'),
            func.sum(Bill.total).label('total_sales')
        ).filter(not_return_clause, Bill.created_at >= start_of_week).first()
        
        # Month's stats
        month_stats = db.session.query(
            func.count(Bill.id).label('bill_count'),
            func.sum(Bill.total).label('total_sales')
        ).filter(not_return_clause, Bill.created_at >= start_of_month).first()
        
        # Pending items count
        pending_items_count = BillItem.query.filter_by(item_status='pending').count()
        
        # Payment method distribution
        payment_methods = db.session.query(
            Bill.payment_method,
            func.count(Bill.id).label('count'),
            func.sum(Bill.total).label('total')
        ).filter(not_return_clause).group_by(Bill.payment_method).all()
        
        # Customer type distribution
        customer_types = db.session.query(
            Bill.customer_type,
            func.count(Bill.id).label('count'),
            func.sum(Bill.total).label('total')
        ).filter(not_return_clause).group_by(Bill.customer_type).all()
        
        # Recent bills
        recent_bills = Bill.query.filter(not_return_clause).order_by(Bill.created_at.desc()).limit(5).all()
        
        return jsonify({
            'today': {
                'bills': today_stats.bill_count or 0,
                'sales': round(today_stats.total_sales or 0, 2),
                'average': round(today_stats.avg_bill_value or 0, 2)
            },
            'thisWeek': {
                'bills': week_stats.bill_count or 0,
                'sales': round(week_stats.total_sales or 0, 2)
            },
            'thisMonth': {
                'bills': month_stats.bill_count or 0,
                'sales': round(month_stats.total_sales or 0, 2)
            },
            'pendingItems': pending_items_count,
            'paymentMethods': [{
                'method': pm[0] or 'other',
                'count': pm[1],
                'total': round(pm[2] or 0, 2)
            } for pm in payment_methods],
            'customerTypes': [{
                'type': ct[0] or 'regular',
                'count': ct[1],
                'total': round(ct[2] or 0, 2)
            } for ct in customer_types],
            'recentBills': [{
                'id': b.id,
                'billNumber': b.bill_number,
                'customerName': b.customer_name,
                'customerType': b.customer_type,
                'vehicleName': b.vehicle_name,
                'vehicleNumber': b.vehicle_number,
                'companyName': b.company_name,
                'total': round(b.total, 2),
                'createdAt': b.created_at.isoformat(),
                'createdBy': b.created_by,
                'createdByName': b.created_by_name
            } for b in recent_bills]
        }), 200
        
    except Exception as e:
        print(f"Statistics error: {str(e)}")
        print(traceback.format_exc())
        return jsonify({"error": "Failed to fetch statistics"}), 400


# ------------------ VOID BILL ITEM ------------------
@billing_bp.route("/billing/bills/<int:bill_id>/items/<int:item_id>/void", methods=["POST"])
def void_bill_item(bill_id, item_id):
    """Void a specific item from bill and adjust stock"""
    try:
        bill = Bill.query.get_or_404(bill_id)
        item = BillItem.query.get_or_404(item_id)
        
        if item.bill_id != bill.id:
            return jsonify({"error": "Item does not belong to this bill"}), 400
        
        # Only restore stock if item is not completed
        if item.item_status != 'completed':
            product = Product.query.get(item.product_id)
            if product:
                product.quantity += item.quantity
        
        # Remove item
        db.session.delete(item)
        
        # Recalculate bill totals
        bill.calculate_totals()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Item voided successfully',
            'bill': bill.to_dict()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Void item error: {str(e)}")
        return jsonify({"error": str(e)}), 400


# ------------------ GET CUSTOMER TYPE SUMMARY ------------------
@billing_bp.route("/billing/customer-types/summary", methods=["GET"])
def get_customer_type_summary():
    """Get summary of bills by customer type"""
    try:
        # Date range parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Build query
        query = db.session.query(
            Bill.customer_type,
            func.count(Bill.id).label('bill_count'),
            func.sum(Bill.total).label('total_sales'),
            func.avg(Bill.total).label('avg_bill_value')
        ).group_by(Bill.customer_type)
        
        if start_date:
            query = query.filter(Bill.created_at >= datetime.fromisoformat(start_date))
        if end_date:
            query = query.filter(Bill.created_at <= datetime.fromisoformat(end_date))
        
        results = query.all()
        
        summary = [{
            'customerType': r[0] or 'regular',
            'billCount': r[1],
            'totalSales': round(r[2] or 0, 2),
            'averageBillValue': round(r[3] or 0, 2)
        } for r in results]
        
        return jsonify({
            'success': True,
            'summary': summary
        }), 200
        
    except Exception as e:
        print(f"Customer type summary error: {str(e)}")
        return jsonify({"error": "Failed to fetch customer type summary"}), 400


# ------------------ GET VEHICLE SUMMARY ------------------
@billing_bp.route("/billing/vehicles/summary", methods=["GET"])
def get_vehicle_summary():
    """Get summary of bills by vehicle number"""
    try:
        # Date range parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Build query for vehicles with bills
        query = db.session.query(
            Bill.vehicle_number,
            Bill.vehicle_name,
            func.count(Bill.id).label('bill_count'),
            func.sum(Bill.total).label('total_spent'),
            func.avg(Bill.total).label('avg_bill_value')
        ).filter(Bill.vehicle_number.isnot(None), Bill.vehicle_number != '')
        
        if start_date:
            query = query.filter(Bill.created_at >= datetime.fromisoformat(start_date))
        if end_date:
            query = query.filter(Bill.created_at <= datetime.fromisoformat(end_date))
        
        results = query.group_by(Bill.vehicle_number, Bill.vehicle_name).order_by(func.sum(Bill.total).desc()).limit(20).all()
        
        vehicles = [{
            'vehicleNumber': r[0],
            'vehicleName': r[1] or '',
            'billCount': r[2],
            'totalSpent': round(r[3] or 0, 2),
            'averageBillValue': round(r[4] or 0, 2)
        } for r in results]
        
        return jsonify({
            'success': True,
            'vehicles': vehicles
        }), 200
        
    except Exception as e:
        print(f"Vehicle summary error: {str(e)}")
        return jsonify({"error": "Failed to fetch vehicle summary"}), 400


# ------------------ GET BILLS BY VEHICLE NUMBER ------------------
@billing_bp.route("/billing/vehicles/<string:vehicle_number>/bills", methods=["GET"])
def get_bills_by_vehicle(vehicle_number):
    """Get all bills for a specific vehicle"""
    try:
        if not vehicle_number:
            return jsonify({"error": "Vehicle number is required"}), 400
        
        bills = Bill.query.filter_by(vehicle_number=vehicle_number).order_by(Bill.created_at.desc()).all()
        
        result = [{
            'id': b.id,
            'billNumber': b.bill_number,
            'customerName': b.customer_name,
            'companyName': b.company_name,
            'total': round(b.total, 2),
            'paidAmount': round(b.paid_amount, 2),
            'paymentStatus': b.payment_status,
            'createdAt': b.created_at.isoformat() if b.created_at else None
        } for b in bills]
        
        return jsonify({
            'success': True,
            'vehicleNumber': vehicle_number,
            'vehicleName': bills[0].vehicle_name if bills else '',
            'bills': result,
            'count': len(result)
        }), 200
        
    except Exception as e:
        print(f"Get bills by vehicle error: {str(e)}")
        return jsonify({"error": "Failed to fetch bills"}), 400
# ==================== WARRANTY ROUTES (Simplified) ====================

# ------------------ WARRANTY SEARCH BY BILL NUMBER ------------------
@billing_bp.route("/billing/warranty/search", methods=["GET"])
def search_warranty_by_bill():
    """Search warranty information by bill number"""
    try:
        bill_number = request.args.get('bill_number')
        
        if not bill_number:
            return jsonify({'error': 'Bill number is required'}), 400
        
        # Use raw SQL to avoid model column issues
        # First, get the bill - using dictionary parameters
        bill_query = """
            SELECT id, bill_number, customer_name, customer_phone, customer_email, 
                   created_at, total 
            FROM bills 
            WHERE bill_number = :bill_number
        """
        bill_result = db.session.execute(text(bill_query), {"bill_number": bill_number})
        bill = bill_result.fetchone()
        
        if not bill:
            return jsonify({'error': 'Bill not found'}), 404
        
        # Get bill items - using dictionary parameters
        items_query = """
            SELECT id, product_id, product_name, product_model, 
                   quantity, sell_price, total 
            FROM bill_items 
            WHERE bill_id = :bill_id
        """
        items_result = db.session.execute(text(items_query), {"bill_id": bill[0]})
        items = items_result.fetchall()
        
        warranty_items = []
        
        for item in items:
            # Get product warranty period from watts field
            product_query = """
                SELECT id, name, model, watts 
                FROM products 
                WHERE id = :product_id
            """
            product_result = db.session.execute(text(product_query), {"product_id": item[1]})
            product = product_result.fetchone()
            
            if not product:
                warranty_period_months = 12  # Default warranty
            else:
                # Get warranty period from watts field (stored in months)
                watts = product[3] if len(product) > 3 else None
                warranty_period_months = int(watts) if watts and watts > 0 else 12
            
            # Warranty start date is bill creation date
            warranty_start_date = bill[5]  # created_at column
            warranty_end_date = warranty_start_date + relativedelta(months=warranty_period_months)
            
            # Calculate warranty status
            current_date = datetime.utcnow()
            
            if current_date <= warranty_end_date:
                days_left = (warranty_end_date - current_date).days
                warranty_status = {
                    'status': 'active',
                    'days_left': days_left,
                    'message': f'Warranty active. {days_left} days remaining'
                }
            else:
                days_expired = (current_date - warranty_end_date).days
                warranty_status = {
                    'status': 'expired',
                    'days_expired': days_expired,
                    'message': f'Warranty expired {days_expired} days ago'
                }
            
            warranty_items.append({
                'productId': item[1],  # product_id
                'productName': item[2],  # product_name
                'productModel': item[3] or 'N/A',  # product_model
                'quantity': item[4],  # quantity
                'sellPrice': float(item[5] or 0),  # sell_price
                'total': float(item[6] or 0),  # total
                'warranty': {
                    'warrantyPeriodMonths': warranty_period_months,
                    'warrantyStartDate': warranty_start_date.isoformat() if warranty_start_date else None,
                    'warrantyEndDate': warranty_end_date.isoformat() if warranty_end_date else None,
                    'warrantyStatus': warranty_status,
                    'isActive': warranty_status['status'] == 'active'
                }
            })
        
        # Bill information
        bill_info = {
            'id': bill[0],
            'billNumber': bill[1],
            'customerName': bill[2] or 'Walk-in Customer',
            'customerPhone': bill[3] or '',
            'customerEmail': bill[4] or '',
            'billedDate': bill[5].isoformat() if bill[5] else None,
            'totalAmount': float(bill[6]) if bill[6] else 0,
            'items': warranty_items
        }
        
        return jsonify(bill_info), 200
        
    except Exception as e:
        print(f"Warranty search error: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


# ------------------ CHECK WARRANTY FOR PRODUCT ------------------
@billing_bp.route("/billing/warranty/check/<int:product_id>/<int:bill_id>", methods=["GET"])
def check_product_warranty(product_id, bill_id):
    """Check warranty status for a specific product in a bill"""
    try:
        # Get bill - using dictionary parameters
        bill_query = """
            SELECT id, bill_number, created_at 
            FROM bills 
            WHERE id = :bill_id
        """
        bill_result = db.session.execute(text(bill_query), {"bill_id": bill_id})
        bill = bill_result.fetchone()
        
        if not bill:
            return jsonify({'error': 'Bill not found'}), 404
        
        # Get product warranty period from watts field
        product_query = """
            SELECT id, name, model, watts 
            FROM products 
            WHERE id = :product_id
        """
        product_result = db.session.execute(text(product_query), {"product_id": product_id})
        product = product_result.fetchone()
        
        if not product:
            return jsonify({'error': 'Product not found'}), 404
        
        # Get warranty period from product's watts field
        watts = product[3] if len(product) > 3 else None
        warranty_period_months = int(watts) if watts and watts > 0 else 12
        
        # Warranty start date is bill creation date
        warranty_start_date = bill[2]
        warranty_end_date = warranty_start_date + relativedelta(months=warranty_period_months)
        
        # Calculate warranty status
        current_date = datetime.utcnow()
        
        if current_date <= warranty_end_date:
            days_left = (warranty_end_date - current_date).days
            warranty_status = {
                'status': 'active',
                'days_left': days_left,
                'message': f'Warranty active. {days_left} days remaining'
            }
        else:
            days_expired = (current_date - warranty_end_date).days
            warranty_status = {
                'status': 'expired',
                'days_expired': days_expired,
                'message': f'Warranty expired {days_expired} days ago'
            }
        
        return jsonify({
            'productId': product[0],
            'productName': product[1],
            'productModel': product[2] or '',
            'billNumber': bill[1],
            'billedDate': warranty_start_date.isoformat(),
            'warrantyPeriodMonths': warranty_period_months,
            'warrantyStartDate': warranty_start_date.isoformat(),
            'warrantyEndDate': warranty_end_date.isoformat(),
            'warrantyStatus': warranty_status
        }), 200
        
    except Exception as e:
        print(f"Check warranty error: {str(e)}")
        return jsonify({'error': str(e)}), 500


# ------------------ MESSENGER / SMS CONFIG ------------------
@billing_bp.route("/billing/messenger-config", methods=["GET"])
def messenger_config():
    """Whether SMS is actually sent to phones or only logged (test mode)."""
    return jsonify(get_messenger_config()), 200


# ------------------ SEND DIGITAL BILL VIA MESSENGER ------------------
@billing_bp.route("/billing/send-digital-bill", methods=["POST"])
def send_digital_bill():
    """Send digital bill link SMS to customer phone (Fast2SMS, Twilio, or mock if unconfigured)."""
    try:
        data = request.get_json() or {}
        raw_phone = str(data.get('phoneNumber', '')).strip()
        customer_name = str(data.get('customerName', 'Customer')).strip() or 'Customer'
        bill_number = str(data.get('billNumber', '')).strip()
        bill_id = data.get('billId') or data.get('bill_id')
        digital_bill_link = str(data.get('digitalBillLink', '')).strip()

        try:
            phone_number = validate_phone_number(raw_phone)
        except SmsValidationError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

        bill = None
        if bill_id is not None:
            try:
                bill = Bill.query.get(int(bill_id))
            except (TypeError, ValueError):
                bill = None
        if not bill and bill_number:
            bill = Bill.query.filter_by(bill_number=bill_number).first()

        if not bill:
            return jsonify({'success': False, 'error': 'Bill not found'}), 404

        bill_number = bill.bill_number
        if not customer_name or customer_name == 'Customer':
            customer_name = bill.customer_name or 'Customer'

        try:
            if not digital_bill_link:
                digital_bill_link = build_digital_bill_url(bill_number)
            message_text = build_digital_bill_sms(customer_name, digital_bill_link)
        except SmsValidationError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

        try:
            message, mode = send_sms(
                phone_number,
                message_text,
                bill_number=bill_number,
            )
        except SmsValidationError as exc:
            return jsonify({'success': False, 'error': str(exc), 'smsDelivered': False}), 400
        except RuntimeError as cfg_err:
            return jsonify({'success': False, 'error': str(cfg_err)}), 503
        except Exception as send_err:
            print(f"[SMS ERROR] {send_err}")
            print(traceback.format_exc())
            return jsonify({
                'success': False,
                'error': f'Failed to send SMS: {send_err}',
            }), 400

        sms_delivered = mode == 'production'

        return jsonify({
            'success': True,
            'smsDelivered': sms_delivered,
            'message': f'Digital bill message sent to {phone_number}',
            'phoneNumber': phone_number,
            'billNumber': bill_number,
            'digitalBillLink': digital_bill_link,
            'messageSid': getattr(message, 'sid', None),
            'messageStatus': getattr(message, 'status', None),
            'mode': mode,
            'provider': mode,
            'timestamp': datetime.now().isoformat(),
        }), 200

    except Exception as e:
        print(f"Send digital bill error: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


# ------------------ GET DIGITAL BILL (Phone Verification) ------------------
@billing_bp.route("/billing/digital-bill/<string:bill_number>/<string:phone_number>", methods=["GET"])
def get_digital_bill(bill_number, phone_number):
    """Get bill details after phone number verification"""
    try:
        # Verify the bill exists and phone number matches
        bill = Bill.query.filter_by(bill_number=bill_number).first()
        
        if not bill:
            return jsonify({'error': 'Bill not found'}), 404

        if not phones_match_bill(bill, phone_number):
            return jsonify({'error': 'Phone number does not match this bill'}), 403

        # Get bill details
        bill_items = BillItem.query.filter_by(bill_id=bill.id).all()

        # Format bill data
        bill_data = {
            'id': bill.id,
            'billNumber': bill.bill_number,
            'customerName': bill.customer_name or 'Walk-in Customer',
            'customerPhone': bill.customer_phone or '',
            'customerAddress': bill.customer_address or '',
            'totalAmount': float(bill.total or 0),
            'subtotal': float(bill.subtotal or 0),
            'discount': float(bill.discount or 0),
            'tax': float(bill.tax or 0),
            'paidAmount': float(bill.paid_amount or 0),
            'paymentMethod': bill.payment_method or 'Cash',
            'createdAt': bill.created_at.isoformat() if bill.created_at else None,
            'items': []
        }

        # Add bill items
        for item in bill_items:
            bill_data['items'].append({
                'productName': item.product_name or '',
                'productModel': item.product_model or '',
                'quantity': item.quantity or 1,
                'price': float(item.sell_price or 0),
                'total': float(item.total or 0)
            })

        return jsonify(bill_data), 200

    except Exception as e:
        print(f"Get digital bill error: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


# ------------------ LEDGER BOOK MULTI-FY REPORT ------------------
def extract_fy_tag(b_str, created_at=None):
    b_str = str(b_str or '').strip().upper()
    import re
    m = re.match(r'^([A-Z]+)\d+[NR]$', b_str)
    if m:
        letters = m.group(1)
        if len(letters) == 1:
            offset = ord(letters[0]) - 65
        else:
            offset = 26 + (ord(letters[-1]) - 65)
        start_yr = (2026 + offset) % 100
        end_yr = (start_yr + 1) % 100
        return f"{start_yr:02d}-{end_yr:02d}"
    elif '/' in b_str:
        return b_str.split('/')[0]
    else:
        return get_current_financial_year(created_at)


@billing_bp.route("/billing/ledger-book", methods=["GET"])
def get_ledger_book_data():
    """Get Multi-FY bills, sales returns, and financial year aggregated summaries for Ledger Book"""
    try:
        selected_fy = request.args.get('fy') # Optional filter e.g. '26-27'
        
        all_bills = Bill.query.order_by(Bill.created_at.desc()).all()
        all_returns = SaleReturn.query.order_by(SaleReturn.created_at.desc()).all()
        
        fy_groups = {}
        processed_bills = []
        
        for bill in all_bills:
            b_str = str(bill.bill_number or '').strip()
            fy_tag = extract_fy_tag(b_str, bill.created_at)
            clean_no = b_str.split('/')[-1] if '/' in b_str else b_str
                
            if fy_tag not in fy_groups:
                fy_groups[fy_tag] = {
                    'fy': fy_tag,
                    'total_bills': 0,
                    'total_sales_amount': 0.0,
                    'total_returns_count': 0,
                    'total_return_amount': 0.0,
                    'net_revenue': 0.0,
                    'total_items_sold': 0
                }
                
            is_return_bill = b_str.upper().endswith('R')
            if not is_return_bill:
                fy_groups[fy_tag]['total_bills'] += 1
                fy_groups[fy_tag]['total_sales_amount'] += float(bill.total or 0)
                qty = sum(int(item.quantity or 1) for item in bill.items) if bill.items else 1
                fy_groups[fy_tag]['total_items_sold'] += qty

            if not selected_fy or selected_fy == fy_tag or selected_fy.lower() == 'all':
                processed_bills.append({
                    'id': bill.id,
                    'fy': fy_tag,
                    'rawBillNumber': bill.bill_number,
                    'billNumber': clean_no,
                    'customerName': bill.customer_name or 'Walk-in Customer',
                    'customerPhone': bill.customer_phone or bill.contact or '',
                    'customerType': bill.customer_type or 'regular',
                    'total': round(float(bill.total or 0), 2),
                    'paidAmount': round(float(bill.paid_amount or 0), 2),
                    'paymentMethod': bill.payment_method or 'cash',
                    'paymentStatus': bill.payment_status or 'completed',
                    'itemCount': len(bill.items),
                    'createdAt': bill.created_at.isoformat() if bill.created_at else None,
                    'createdByName': bill.created_by_name or 'Admin'
                })

        for ret in all_returns:
            r_str = str(ret.return_number or '').strip()
            fy_tag = extract_fy_tag(r_str, ret.created_at)
                
            if fy_tag not in fy_groups:
                fy_groups[fy_tag] = {
                    'fy': fy_tag,
                    'total_bills': 0,
                    'total_sales_amount': 0.0,
                    'total_returns_count': 0,
                    'total_return_amount': 0.0,
                    'net_revenue': 0.0,
                    'total_items_sold': 0
                }
            fy_groups[fy_tag]['total_returns_count'] += 1
            fy_groups[fy_tag]['total_return_amount'] += float(ret.total_return_amount or 0)

        for fy_key in fy_groups:
            g = fy_groups[fy_key]
            g['total_sales_amount'] = round(g['total_sales_amount'], 2)
            g['total_return_amount'] = round(g['total_return_amount'], 2)
            g['net_revenue'] = round(g['total_sales_amount'] - g['total_return_amount'], 2)

        available_fy = sorted(list(fy_groups.keys()), reverse=True)
        if not available_fy:
            curr_fy = get_current_financial_year()
            available_fy = [curr_fy]
            fy_groups[curr_fy] = {
                'fy': curr_fy,
                'total_bills': 0,
                'total_sales_amount': 0.0,
                'total_returns_count': 0,
                'total_return_amount': 0.0,
                'net_revenue': 0.0,
                'total_items_sold': 0
            }

        return jsonify({
            'success': True,
            'available_fy_years': available_fy,
            'fy_summaries': [fy_groups[k] for k in sorted(fy_groups.keys(), reverse=True)],
            'bills': processed_bills
        }), 200

    except Exception as e:
        print(f"Ledger book error: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

